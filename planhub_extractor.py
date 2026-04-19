"""
PlanHub Extractor — Tesseract OCR  (LEAN VERSION)
==================================================
Extracts only what's reliable from PlanHub profile screenshots:
  • Company name
  • Email
  • Phone
  • Website
  • Address (street / city / state)

Fast (~2-3s per screenshot on dual-core), offline, no LLM.

SETUP (one-time per laptop):
  Option A (recommended):  Distribute the bundled .exe — no installs needed
  Option B (from source):
      1. Install Tesseract: https://github.com/UB-Mannheim/tesseract/wiki
      2. pip install pytesseract pillow openpyxl tqdm

USAGE:
  Run the .exe (or: python planhub_extractor.py)
  Paste the screenshots folder path. Excel appears in the same folder.
"""

import json
import os
import re
import shutil
import sys
import threading
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import pytesseract
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from tqdm import tqdm

# ──────────── CONFIG ────────────────────────────────────────────────────────
WORKERS = 2            # Dual-core safe. 2 workers = ~300MB RAM total.
DONE_FOLDER = "done"
# ────────────────────────────────────────────────────────────────────────────

# Tesseract path autodetect (Windows)
if sys.platform == "win32":
    common_paths = [
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    ]
    for p in common_paths:
        if Path(p).exists():
            pytesseract.pytesseract.tesseract_cmd = p
            break
    # PyInstaller bundled path (when shipped as .exe)
    if hasattr(sys, "_MEIPASS"):
        bundled = Path(sys._MEIPASS) / "tesseract" / "tesseract.exe"
        if bundled.exists():
            pytesseract.pytesseract.tesseract_cmd = str(bundled)
            os.environ["TESSDATA_PREFIX"] = str(Path(sys._MEIPASS) / "tesseract" / "tessdata")

FIELDS = ["company_name", "email", "phone", "website", "address", "city", "state"]
COLUMNS = FIELDS + ["source_file"]
HEADER_LABELS = ["Company Name", "Email", "Phone", "Website",
                 "Address", "City", "State", "Source File"]

NAV_WORDS = {
    "planhub", "planhiub", "plantiub",
    "projects", "lead", "finder", "takeoff", "estimation",
    "bid", "planner", "bids", "directory", "messages", "company",
    "request", "demo", "free", "trial", "upgrade", "days", "left", "miles", "from",
}
BADGE_WORDS = {"premium", "subcontractor", "verified"}

save_lock = threading.Lock()


# ═══════════════════════════════════════════════════════════════════════════
# OCR HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def get_words(img: Image.Image) -> list:
    """Word-level OCR with bounding boxes."""
    data = pytesseract.image_to_data(img, output_type=pytesseract.Output.DICT)
    out = []
    for i in range(len(data["text"])):
        text = data["text"][i].strip()
        if not text:
            continue
        out.append({
            "text": text,
            "x": data["left"][i],
            "y": data["top"][i],
            "w": data["width"][i],
            "h": data["height"][i],
            "cx": data["left"][i] + data["width"][i] // 2,
            "cy": data["top"][i] + data["height"][i] // 2,
            "line": (data["block_num"][i], data["par_num"][i], data["line_num"][i]),
        })
    return out


def group_lines(words: list) -> list:
    by_line = defaultdict(list)
    for w in words:
        by_line[w["line"]].append(w)
    result = []
    for ws in by_line.values():
        ws = sorted(ws, key=lambda w: w["x"])
        result.append({
            "text": " ".join(w["text"] for w in ws),
            "y": min(w["y"] for w in ws),
            "cy": sum(w["cy"] for w in ws) // len(ws),
            "words": ws,
        })
    result.sort(key=lambda l: l["y"])
    return result


def strip_icon_prefix(text: str) -> str:
    """Remove map-pin icon OCR-artifacts ('Q ', '@ ', '2 ', etc.) before address."""
    return re.sub(r"^[Q@©®¢9¥2§°]+\s+", "", text).strip()


# ═══════════════════════════════════════════════════════════════════════════
# FIELD EXTRACTION
# ═══════════════════════════════════════════════════════════════════════════

def parse_header(img: Image.Image) -> dict:
    """Company name + address from the dark header band."""
    W, H = img.size
    crop = img.crop((0, int(H * 0.08), int(W * 0.68), int(H * 0.32)))
    words = get_words(crop)
    lines = group_lines(words)
    result = {"company_name": "", "address": "", "city": "", "state": ""}

    # Company name
    for line in lines:
        text = line["text"].strip()
        if not text:
            continue
        wlower = {re.sub(r"[^a-z]", "", w.lower()) for w in text.split()}
        if len(wlower & NAV_WORDS) >= 2:
            continue
        if re.search(r",\s*[A-Z]{2}\b", text):
            continue
        if re.search(r"^[^A-Za-z]*trades\s*:", text, re.I):
            continue
        tokens = text.split()
        while tokens and re.sub(r"[^a-z]", "", tokens[0].lower()) in BADGE_WORDS:
            tokens.pop(0)
        while tokens and not re.search(r"[A-Za-z0-9]", tokens[0]):
            tokens.pop(0)
        while tokens:
            last = re.sub(r"[^a-z]", "", tokens[-1].lower())
            if last in {"", "gc", "cc", "ce", "itbs", "insig", "ecc", "occ"}:
                tokens.pop()
            else:
                break
        cleaned = " ".join(tokens)
        cleaned = re.sub(r"\s+[@©®]+(\s+[@©®]+)*$", "", cleaned).strip()
        if len(cleaned.split()) == 1 and cleaned.isupper() and len(cleaned) <= 4:
            continue
        if cleaned.lower() in BADGE_WORDS:
            continue
        if len(cleaned) >= 3 and re.search(r"[A-Za-z]{2,}", cleaned):
            result["company_name"] = cleaned
            break

    # Address
    for line in lines:
        clean = strip_icon_prefix(line["text"])
        m = re.match(r"^(.+?),\s*(.+?),\s*([A-Za-z]{2})\b", clean)
        if m:
            result["address"] = m.group(1).strip()
            result["city"] = m.group(2).strip()
            result["state"] = m.group(3).strip().upper()
            break

    return result


def parse_general_info(img: Image.Image) -> dict:
    """Website, email, phone from General Information section."""
    W, H = img.size
    crop = img.crop((0, int(H * 0.30), int(W * 0.68), int(H * 0.55)))
    words = get_words(crop)
    cw = crop.width
    result = {"website": "", "email": "", "phone": ""}

    def find_label(tokens: list):
        lower_tokens = [t.lower() for t in tokens]
        for i, w in enumerate(words):
            if w["text"].lower().rstrip(":") != lower_tokens[0]:
                continue
            matched = [w]
            ok = True
            for j, tok in enumerate(lower_tokens[1:], 1):
                if i + j >= len(words):
                    ok = False
                    break
                nxt = words[i + j]
                if nxt["text"].lower().rstrip(":") != tok or abs(nxt["cy"] - w["cy"]) > 12:
                    ok = False
                    break
                matched.append(nxt)
            if ok:
                return matched
        return None

    def value_right_of(match, x_limit):
        if not match:
            return ""
        last = match[-1]
        y = last["cy"]
        vals = [w for w in words
                if w not in match
                and abs(w["cy"] - y) < 12
                and w["x"] > last["x"] + last["w"]
                and w["x"] < x_limit]
        vals.sort(key=lambda w: w["x"])
        return " ".join(v["text"] for v in vals).strip()

    left_cap = int(cw * 0.55)

    # Website
    val = value_right_of(find_label(["Website"]), left_cap)
    if val:
        m = re.search(
            r"(https?://\S+|www\.\S+|[\w.-]+\.(?:com|net|org|io|us|co|biz|info|tv)(?:/\S*)?)",
            val, re.I)
        if m:
            result["website"] = m.group(1).rstrip(".,;")

    # Email
    val = value_right_of(find_label(["Email"]), left_cap)
    if val:
        m = re.search(r"[\w.+-]+@[\w.-]+\.[a-zA-Z]{2,}", val)
        if m:
            result["email"] = m.group(0)

    # Phone
    val = value_right_of(find_label(["Telephone", "Number"]), left_cap)
    if val:
        digits = re.sub(r"\D", "", val)
        if len(digits) == 10:
            result["phone"] = digits
        elif len(digits) == 11 and digits.startswith("1"):
            result["phone"] = digits[1:]

    # Safety-net fallback: scan the whole crop's text for any missed field.
    # This catches edge cases where column alignment got messed up.
    if not result["email"] or not result["website"] or not result["phone"]:
        full_text = " ".join(w["text"] for w in words)
        if not result["email"]:
            m = re.search(r"[\w.+-]+@[\w.-]+\.[a-zA-Z]{2,}", full_text)
            if m:
                result["email"] = m.group(0)
        if not result["website"]:
            m = re.search(
                r"(https?://\S+|www\.[\w.-]+\.[a-z]{2,}(?:/\S*)?)",
                full_text, re.I)
            if m:
                result["website"] = m.group(1).rstrip(".,;")
        if not result["phone"]:
            m = re.search(r"\(?\d{3}\)?[\s.-]?\d{3}[\s.-]?\d{4}", full_text)
            if m:
                digits = re.sub(r"\D", "", m.group(0))
                if len(digits) == 10:
                    result["phone"] = digits

    return result


def extract_all(image_path: Path) -> dict:
    img = Image.open(image_path)
    if img.mode != "RGB":
        img = img.convert("RGB")

    record = {f: "" for f in FIELDS}
    record["source_file"] = image_path.name

    record.update(parse_header(img))
    record.update(parse_general_info(img))
    return record


# ═══════════════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ═══════════════════════════════════════════════════════════════════════════

def save_excel(records: list, out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "PlanHub Data"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    side = Side(style="thin", color="CCCCCC")
    thin = Border(left=side, right=side, top=side, bottom=side)

    for c, label in enumerate(HEADER_LABELS, start=1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws.row_dimensions[1].height = 30

    for r, rec in enumerate(records, start=2):
        fill = alt_fill if r % 2 == 0 else PatternFill()
        for c, field in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r, column=c, value=rec.get(field, ""))
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin

    widths = [30, 32, 14, 32, 32, 20, 8, 30]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(out_path)


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  PlanHub Extractor  —  Tesseract OCR  (offline)")
    print("=" * 60)

    try:
        ver = pytesseract.get_tesseract_version()
        print(f"  ✅ Tesseract {ver} ready  |  Workers: {WORKERS}")
    except Exception:
        print("\n❌ Tesseract not found!")
        if sys.platform == "win32":
            print("   Install: https://github.com/UB-Mannheim/tesseract/wiki")
            print("   (Default path: C:\\Program Files\\Tesseract-OCR\\)")
        else:
            print("   Install: sudo apt install tesseract-ocr")
        input("\nPress Enter to exit...")
        sys.exit(1)
    print("=" * 60)

    folder_raw = input("\nPaste your screenshots folder path and press Enter:\n> ").strip().strip('"')
    folder = Path(folder_raw)
    if not folder.exists():
        print(f"❌ Folder not found: {folder}")
        input("\nPress Enter to exit...")
        sys.exit(1)

    done_folder = folder / DONE_FOLDER
    done_folder.mkdir(exist_ok=True)

    exts = {".png", ".jpg", ".jpeg", ".webp", ".bmp"}
    all_images = sorted([f for f in folder.iterdir()
                         if f.is_file() and f.suffix.lower() in exts])
    if not all_images:
        print("❌ No image files found.")
        input("\nPress Enter to exit...")
        sys.exit(1)

    progress_file = folder / "planhub_progress.json"
    output_excel = folder / "planhub_data.xlsx"

    done_files = {}
    if progress_file.exists():
        try:
            with open(progress_file, "r", encoding="utf-8") as f:
                done_files = json.load(f)
            print(f"\n📂 Resuming — {len(done_files)} already done.")
        except Exception:
            pass

    to_process = [f for f in all_images if f.name not in done_files]
    total = len(all_images)
    remaining = len(to_process)

    secs_est = (remaining * 3) // max(WORKERS, 1)
    mins = secs_est // 60
    time_str = f"~{mins} min" if mins > 0 else f"~{secs_est} sec"

    print(f"\n📸 Total screenshots : {total}")
    print(f"✅ Already processed : {total - remaining}")
    print(f"🔄 To process now    : {remaining}")
    print(f"⚡ Workers           : {WORKERS}")
    print(f"⏱️  Estimated time    : {time_str}")
    print(f"📊 Output            : {output_excel}")
    print(f"📁 Done folder       : {done_folder}")
    print("\n💡 Saves after every image. Ctrl+C anytime to pause.\n")
    input("Press Enter to start...")

    pbar = tqdm(total=remaining, unit="img", dynamic_ncols=True)

    def process_one(img_path: Path):
        try:
            record = extract_all(img_path)
        except Exception as e:
            tqdm.write(f"  ❌ Error on {img_path.name}: {e}")
            record = {f: "" for f in FIELDS}
            record["source_file"] = img_path.name
        with save_lock:
            done_files[img_path.name] = record
            with open(progress_file, "w", encoding="utf-8") as f:
                json.dump(done_files, f, ensure_ascii=False, indent=2)
            save_excel(list(done_files.values()), output_excel)
            dest = done_folder / img_path.name
            if dest.exists():
                dest = done_folder / f"{img_path.stem}_dup{img_path.suffix}"
            try:
                shutil.move(str(img_path), str(dest))
            except Exception as e:
                tqdm.write(f"  ⚠️  Move failed for {img_path.name}: {e}")
        pbar.set_description(img_path.name[:35])
        pbar.update(1)

    try:
        with ThreadPoolExecutor(max_workers=WORKERS) as executor:
            futures = {executor.submit(process_one, img): img for img in to_process}
            for future in as_completed(futures):
                try:
                    future.result()
                except Exception as e:
                    tqdm.write(f"  ❌ Worker error: {e}")
    except KeyboardInterrupt:
        print("\n\n⏸️  Paused. Re-run to resume.")
        pbar.close()
        sys.exit(0)

    pbar.close()
    print(f"\n✅ Done! {len(done_files)} rows → {output_excel}")
    if progress_file.exists():
        progress_file.unlink()
    print("\n🎉 Finished!")
    input("\nPress Enter to exit...")


if __name__ == "__main__":
    main()
